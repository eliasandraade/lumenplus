"""
Export Routes
=============
Endpoints de exportação de dados de usuários com dupla confirmação para dados sensíveis.

Fluxo:
- Sem RG/CPF: CSV gerado imediatamente (status GENERATED)
- Com RG/CPF:  status PENDING → aprovação por COUNCIL_GENERAL/DEV/ADMIN → GENERATED
               Link disponível por 24h, depois EXPIRED
"""

import csv
import io
import os
from datetime import datetime, timedelta, timezone
from typing import Any, Literal
from uuid import UUID

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from sqlalchemy import select

from app.api.deps import CurrentUser, DBSession
from app.db.models import (
    AuditLog,
    DataExportRequest,
    GlobalRole,
    User,
    UserGlobalRole,
    UserIdentity,
    UserProfile,
    ProfileCatalogItem,
)
from app.services.organization import get_user_global_roles

router = APIRouter(prefix="/admin/export", tags=["Export"])

SENSITIVE_FIELDS = {"cpf", "rg"}
ALLOWED_FIELDS = {
    "name", "email", "phone", "city", "state", "birth_date",
    "instagram", "profile_status", "cpf", "rg",
    "realidade_vocacional", "estado_civil", "estado_de_vida",
    "acompanhamento_vocacional", "interesse_ministerio",
    "consagracao_ano", "global_roles",
}
EXPORT_TTL_HOURS = 24


def _require_export_permission(db, user_id: UUID) -> None:
    roles = get_user_global_roles(db, user_id)
    if not any(r in roles for r in ["DEV", "ADMIN", "SECRETARY"]):
        raise HTTPException(status_code=403, detail={"error": "forbidden"})


def _require_approval_permission(db, user_id: UUID) -> None:
    roles = get_user_global_roles(db, user_id)
    if not any(r in roles for r in ["DEV", "ADMIN", "COUNCIL_GENERAL"]):
        raise HTTPException(status_code=403, detail={"error": "forbidden"})


def _catalog_label(db, item_id) -> str:
    """Retorna o label de um ProfileCatalogItem pelo ID."""
    if not item_id:
        return ""
    item = db.execute(
        select(ProfileCatalogItem).where(ProfileCatalogItem.id == item_id)
    ).scalar_one_or_none()
    return item.label if item else ""


def _generate_csv(db, fields: list[str], filters: dict) -> str:
    """Gera CSV em memória com os usuários."""
    from app.crypto.service import crypto_service
    from app.services.organization import get_user_global_roles as _get_roles
    crypto = crypto_service

    stmt = (
        select(User)
        .outerjoin(UserProfile, User.id == UserProfile.user_id)
        .outerjoin(UserIdentity, User.id == UserIdentity.user_id)
    )
    users = db.execute(stmt).scalars().unique().all()

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fields)
    writer.writeheader()

    for user in users:
        profile = user.profile
        identity = user.identities[0] if user.identities else None
        row = {}
        for field in fields:
            if field == "name":
                row["name"] = profile.full_name if profile else ""
            elif field == "email":
                row["email"] = identity.email if identity else ""
            elif field == "phone":
                row["phone"] = profile.phone_e164 if profile else ""
            elif field == "city":
                row["city"] = profile.city if profile else ""
            elif field == "state":
                row["state"] = profile.state if profile else ""
            elif field == "birth_date":
                row["birth_date"] = profile.birth_date.isoformat() if profile and profile.birth_date else ""
            elif field == "instagram":
                row["instagram"] = profile.instagram if profile else ""
            elif field == "profile_status":
                row["profile_status"] = profile.status if profile else "INCOMPLETE"
            elif field == "realidade_vocacional":
                row["realidade_vocacional"] = _catalog_label(db, profile.vocational_reality_item_id if profile else None)
            elif field == "estado_civil":
                row["estado_civil"] = _catalog_label(db, profile.marital_status_item_id if profile else None)
            elif field == "estado_de_vida":
                row["estado_de_vida"] = _catalog_label(db, profile.life_state_item_id if profile else None)
            elif field == "acompanhamento_vocacional":
                row["acompanhamento_vocacional"] = (
                    "Sim" if profile and profile.has_vocational_accompaniment else
                    "Não" if profile and profile.has_vocational_accompaniment is not None else ""
                )
            elif field == "interesse_ministerio":
                row["interesse_ministerio"] = (
                    "Sim" if profile and profile.interested_in_ministry else
                    "Não" if profile and profile.interested_in_ministry is not None else ""
                )
            elif field == "consagracao_ano":
                row["consagracao_ano"] = str(profile.consecration_year) if profile and profile.consecration_year else ""
            elif field == "global_roles":
                roles = _get_roles(db, user.id)
                row["global_roles"] = ", ".join(roles)
            elif field == "cpf":
                if profile and profile.cpf_encrypted:
                    try:
                        row["cpf"] = crypto.decrypt(profile.cpf_encrypted)
                    except Exception:
                        row["cpf"] = ""
                else:
                    row["cpf"] = ""
            elif field == "rg":
                if profile and profile.rg_encrypted:
                    try:
                        row["rg"] = crypto.decrypt(profile.rg_encrypted)
                    except Exception:
                        row["rg"] = ""
                else:
                    row["rg"] = ""
        writer.writerow(row)

    return output.getvalue()


# Cabeçalhos legíveis em português para o XLSX
_FIELD_LABELS: dict[str, str] = {
    "name":                      "Nome Completo",
    "email":                     "E-mail",
    "phone":                     "Telefone",
    "birth_date":                "Data de Nascimento",
    "city":                      "Cidade",
    "state":                     "Estado (UF)",
    "instagram":                 "Instagram",
    "profile_status":            "Status do Perfil",
    "realidade_vocacional":      "Realidade Vocacional",
    "estado_civil":              "Estado Civil",
    "estado_de_vida":            "Estado de Vida",
    "acompanhamento_vocacional": "Acompanhamento Vocacional",
    "interesse_ministerio":      "Interesse em Ministério",
    "consagracao_ano":           "Ano de Consagração",
    "global_roles":              "Cargos",
    "cpf":                       "CPF",
    "rg":                        "RG",
}


def _generate_xlsx(db, fields: list[str], filters: dict) -> bytes:
    """Gera XLSX em memória com cabeçalhos formatados e colunas ajustadas."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Reutiliza a lógica de geração de dados do CSV (sem serializar)
    from app.crypto.service import crypto_service
    from app.services.organization import get_user_global_roles as _get_roles

    crypto = crypto_service

    stmt = (
        select(User)
        .outerjoin(UserProfile, User.id == UserProfile.user_id)
        .outerjoin(UserIdentity, User.id == UserIdentity.user_id)
    )
    users = db.execute(stmt).scalars().unique().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuários Lumen+"

    # ── Cabeçalho estilizado ─────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="7C3AED")   # roxo admin
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="E8E8E8")
    border = Border(left=thin, right=thin, bottom=thin)

    for col_idx, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_FIELD_LABELS.get(field, field))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 28

    # ── Dados ────────────────────────────────────────────────────────────────
    data_align = Alignment(vertical="center")
    even_fill = PatternFill("solid", fgColor="F9F5FF")     # roxo muito claro

    for row_idx, user in enumerate(users, start=2):
        profile = user.profile
        identity = user.identities[0] if user.identities else None
        fill = even_fill if row_idx % 2 == 0 else None

        for col_idx, field in enumerate(fields, start=1):
            value: Any = ""
            if field == "name":
                value = profile.full_name if profile else ""
            elif field == "email":
                value = identity.email if identity else ""
            elif field == "phone":
                value = profile.phone_e164 if profile else ""
            elif field == "city":
                value = profile.city if profile else ""
            elif field == "state":
                value = profile.state if profile else ""
            elif field == "birth_date":
                value = profile.birth_date if profile and profile.birth_date else ""
            elif field == "instagram":
                value = profile.instagram if profile else ""
            elif field == "profile_status":
                raw = profile.status if profile else "INCOMPLETE"
                value = "Completo" if raw == "COMPLETE" else "Incompleto"
            elif field == "realidade_vocacional":
                value = _catalog_label(db, profile.vocational_reality_item_id if profile else None)
            elif field == "estado_civil":
                value = _catalog_label(db, profile.marital_status_item_id if profile else None)
            elif field == "estado_de_vida":
                value = _catalog_label(db, profile.life_state_item_id if profile else None)
            elif field == "acompanhamento_vocacional":
                value = (
                    "Sim" if profile and profile.has_vocational_accompaniment else
                    "Não" if profile and profile.has_vocational_accompaniment is not None else ""
                )
            elif field == "interesse_ministerio":
                value = (
                    "Sim" if profile and profile.interested_in_ministry else
                    "Não" if profile and profile.interested_in_ministry is not None else ""
                )
            elif field == "consagracao_ano":
                value = profile.consecration_year if profile and profile.consecration_year else ""
            elif field == "global_roles":
                value = ", ".join(_get_roles(db, user.id))
            elif field == "cpf":
                if profile and profile.cpf_encrypted:
                    try:
                        value = crypto.decrypt(profile.cpf_encrypted)
                    except Exception:
                        value = ""
            elif field == "rg":
                if profile and profile.rg_encrypted:
                    try:
                        value = crypto.decrypt(profile.rg_encrypted)
                    except Exception:
                        value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
            cell.border = border
            if fill:
                cell.fill = fill

        ws.row_dimensions[row_idx].height = 20

    # ── Ajuste automático de largura das colunas ─────────────────────────────
    for col_idx, field in enumerate(fields, start=1):
        col_letter = get_column_letter(col_idx)
        header_len = len(_FIELD_LABELS.get(field, field))
        max_len = max(
            header_len,
            *[
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(2, min(ws.max_row + 1, 102))  # sample primeiras 100 linhas
            ],
            1,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    # Congela a linha de cabeçalho
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ExportRequestBody(BaseModel):
    fields: list[str]
    filters: dict = {}
    format: Literal["csv", "xlsx"] = "csv"


@router.post("/request")
async def create_export_request(
    body: ExportRequestBody,
    current_user: CurrentUser,
    db: DBSession,
) -> Any:
    """Solicita exportação. CSV imediato se sem dados sensíveis, PENDING se tiver."""
    _require_export_permission(db, current_user.id)

    invalid = set(body.fields) - ALLOWED_FIELDS
    if invalid:
        raise HTTPException(
            status_code=422,
            detail={"error": "invalid_fields", "fields": list(invalid)},
        )

    has_sensitive = bool(set(body.fields) & SENSITIVE_FIELDS)

    export_req = DataExportRequest(
        requested_by=current_user.id,
        status="PENDING",
        fields_requested=body.fields,
        filters_json=body.filters,
        has_sensitive=has_sensitive,
    )
    db.add(export_req)
    db.flush()

    db.add(
        AuditLog(
            actor_user_id=current_user.id,
            action="EXPORT_REQUESTED",
            entity_type="DATA_EXPORT",
            entity_id=str(export_req.id),
            extra_data={"fields": body.fields, "has_sensitive": has_sensitive},
        )
    )

    # Também salva o formato solicitado para uso no download posterior
    fmt = body.format  # "csv" ou "xlsx"

    if not has_sensitive:
        export_req.status = "GENERATED"
        export_req.expires_at = datetime.now(timezone.utc) + timedelta(hours=EXPORT_TTL_HOURS)
        # Guarda o formato no filters_json para reusar no download
        if export_req.filters_json is None:
            export_req.filters_json = {}
        export_req.filters_json = {**(export_req.filters_json or {}), "_fmt": fmt}
        db.commit()

        if fmt == "xlsx":
            content = _generate_xlsx(db, body.fields, body.filters)
            filename = f"lumenplus_usuarios_{export_req.id}.xlsx"
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "X-Export-Id": str(export_req.id),
                },
            )
        else:
            csv_content = _generate_csv(db, body.fields, body.filters)
            filename = f"lumenplus_usuarios_{export_req.id}.csv"
            return Response(
                content=csv_content.encode("utf-8-sig"),
                media_type="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "X-Export-Id": str(export_req.id),
                },
            )
    else:
        _notify_council_for_approval(db, export_req, current_user.id)
        db.commit()
        return {
            "id": str(export_req.id),
            "status": export_req.status,
            "has_sensitive": has_sensitive,
            "message": "Aguardando aprovação do Conselho Geral",
        }


def _notify_council_for_approval(db, export_req: DataExportRequest, requester_id: UUID) -> None:
    """Envia mensagem no Inbox para todos os usuários com cargo COUNCIL_GENERAL."""
    requester = db.execute(select(User).where(User.id == requester_id)).scalar_one_or_none()
    requester_name = requester.profile.full_name if requester and requester.profile else "Alguém"
    fields_str = ", ".join(export_req.fields_requested)

    council_user_ids = db.execute(
        select(UserGlobalRole.user_id)
        .join(GlobalRole)
        .where(GlobalRole.code == "COUNCIL_GENERAL")
    ).scalars().all()

    if not council_user_ids:
        return

    from app.db.models import InboxApprovalStatus, InboxMessage, InboxMessageType, InboxRecipient

    inbox_msg = InboxMessage(
        title="Aprovação necessária: Exportação de dados sensíveis",
        message=(
            f"{requester_name} solicitou uma exportação de dados que inclui informações sensíveis "
            f"(campos: {fields_str}). Acesse a área de Aprovações no painel admin para aprovar ou rejeitar."
        ),
        type=InboxMessageType.INFO,
        created_by_user_id=requester_id,
        expires_at=datetime.now(timezone.utc) + timedelta(days=7),
        approval_status=InboxApprovalStatus.AUTO_APPROVED,
    )
    db.add(inbox_msg)
    db.flush()
    for uid in council_user_ids:
        db.add(InboxRecipient(message_id=inbox_msg.id, user_id=uid))


@router.get("/requests")
async def list_export_requests(
    current_user: CurrentUser,
    db: DBSession,
) -> Any:
    """Lista solicitações. DEV/ADMIN/COUNCIL_GENERAL veem todas; outros veem as próprias."""
    caller_roles = get_user_global_roles(db, current_user.id)
    can_see_all = any(r in caller_roles for r in ["DEV", "ADMIN", "COUNCIL_GENERAL"])

    stmt = select(DataExportRequest).order_by(DataExportRequest.created_at.desc())
    if not can_see_all:
        stmt = stmt.where(DataExportRequest.requested_by == current_user.id)

    reqs = db.execute(stmt).scalars().all()
    return [
        {
            "id": str(r.id),
            "requested_by": str(r.requested_by),
            "status": r.status,
            "fields_requested": r.fields_requested,
            "has_sensitive": r.has_sensitive,
            "approved_by": str(r.approved_by) if r.approved_by else None,
            "approved_at": r.approved_at.isoformat() if r.approved_at else None,
            "expires_at": r.expires_at.isoformat() if r.expires_at else None,
            "created_at": r.created_at.isoformat(),
        }
        for r in reqs
    ]


@router.post("/{export_id}/approve")
async def approve_export(
    export_id: UUID,
    current_user: CurrentUser,
    db: DBSession,
) -> Any:
    """Aprova exportação pendente. Requer COUNCIL_GENERAL, DEV ou ADMIN."""
    _require_approval_permission(db, current_user.id)

    export_req = db.execute(
        select(DataExportRequest).where(DataExportRequest.id == export_id)
    ).scalar_one_or_none()
    if not export_req:
        raise HTTPException(status_code=404, detail={"error": "not_found"})
    if export_req.status != "PENDING":
        raise HTTPException(
            status_code=409,
            detail={"error": "not_pending", "current_status": export_req.status},
        )

    now = datetime.now(timezone.utc)
    export_req.status = "GENERATED"
    export_req.approved_by = current_user.id
    export_req.approved_at = now
    export_req.expires_at = now + timedelta(hours=EXPORT_TTL_HOURS)

    db.add(
        AuditLog(
            actor_user_id=current_user.id,
            action="EXPORT_APPROVED",
            entity_type="DATA_EXPORT",
            entity_id=str(export_id),
        )
    )

    _notify_requester_approved(db, export_req, current_user.id)
    db.commit()
    return {"id": str(export_req.id), "status": export_req.status}


def _notify_requester_approved(db, export_req: DataExportRequest, approver_id: UUID) -> None:
    from app.db.models import InboxApprovalStatus, InboxMessage, InboxMessageType, InboxRecipient

    inbox_msg = InboxMessage(
        title="Exportação aprovada",
        message="Sua exportação de dados foi aprovada e está disponível para download por 24 horas.",
        type=InboxMessageType.INFO,
        created_by_user_id=approver_id,
        expires_at=datetime.now(timezone.utc) + timedelta(days=2),
        approval_status=InboxApprovalStatus.AUTO_APPROVED,
    )
    db.add(inbox_msg)
    db.flush()
    db.add(InboxRecipient(message_id=inbox_msg.id, user_id=export_req.requested_by))


@router.post("/{export_id}/reject")
async def reject_export(
    export_id: UUID,
    current_user: CurrentUser,
    db: DBSession,
) -> Any:
    """Rejeita exportação pendente."""
    _require_approval_permission(db, current_user.id)

    export_req = db.execute(
        select(DataExportRequest).where(DataExportRequest.id == export_id)
    ).scalar_one_or_none()
    if not export_req:
        raise HTTPException(status_code=404, detail={"error": "not_found"})
    if export_req.status != "PENDING":
        raise HTTPException(status_code=409, detail={"error": "not_pending"})

    export_req.status = "REJECTED"
    db.add(
        AuditLog(
            actor_user_id=current_user.id,
            action="EXPORT_REJECTED",
            entity_type="DATA_EXPORT",
            entity_id=str(export_id),
        )
    )
    db.commit()
    return {"id": str(export_req.id), "status": "REJECTED"}


@router.get("/{export_id}/download")
async def download_export(
    export_id: UUID,
    current_user: CurrentUser,
    db: DBSession,
) -> Response:
    """Baixa o CSV gerado na hora. Apenas solicitante ou admins. Registra auditoria."""
    caller_roles = get_user_global_roles(db, current_user.id)
    is_admin = any(r in caller_roles for r in ["DEV", "ADMIN"])

    export_req = db.execute(
        select(DataExportRequest).where(DataExportRequest.id == export_id)
    ).scalar_one_or_none()
    if not export_req:
        raise HTTPException(status_code=404, detail={"error": "not_found"})

    if export_req.requested_by != current_user.id and not is_admin:
        raise HTTPException(status_code=403, detail={"error": "forbidden"})

    if export_req.status != "GENERATED":
        raise HTTPException(
            status_code=409,
            detail={"error": "not_ready", "status": export_req.status},
        )

    now = datetime.now(timezone.utc)
    if export_req.expires_at and now > export_req.expires_at:
        export_req.status = "EXPIRED"
        db.commit()
        raise HTTPException(status_code=410, detail={"error": "expired"})

    # Descobre o formato solicitado originalmente (_fmt salvo em filters_json)
    filters = dict(export_req.filters_json or {})
    fmt = filters.pop("_fmt", "csv")  # remove a chave interna antes de passar aos filtros

    db.add(
        AuditLog(
            actor_user_id=current_user.id,
            action="EXPORT_DOWNLOADED",
            entity_type="DATA_EXPORT",
            entity_id=str(export_id),
        )
    )
    db.commit()

    if fmt == "xlsx":
        content = _generate_xlsx(db, export_req.fields_requested, filters)
        filename = f"lumenplus_usuarios_{export_id}.xlsx"
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # CSV (padrão)
    csv_content = _generate_csv(db, export_req.fields_requested, filters)
    filename = f"lumenplus_usuarios_{export_id}.csv"
    return Response(
        content=csv_content.encode("utf-8-sig"),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
